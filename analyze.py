"""
分析歷史採樣資料，找出「過去 X 小時 sbi 一直是 1 且 mday 都相同」的站
產出：
  - stale.json （給 HTML 讀）
  - stale.xlsx （給同事下載）
"""
import json
import os
import glob
import sys
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

TPE = timezone(timedelta(hours=8))
HISTORY_DIR = "history"
WINDOWS = [1, 2, 4]         # 要分析的小時區間
TARGET_SBI = 1              # 「只有 1 台車」


def load_history():
    """讀出所有採樣檔，回傳 [(sampled_at, {sno: station_data}), ...]，依時間排序"""
    samples = []
    for f in sorted(glob.glob(os.path.join(HISTORY_DIR, "*.json"))):
        try:
            with open(f, "r", encoding="utf-8") as fp:
                data = json.load(fp)
            t = datetime.fromisoformat(data["sampled_at"])
            samples.append((t, data["stations"]))
        except Exception as e:
            print(f"⚠ 讀 {f} 失敗: {e}", file=sys.stderr)
    return samples


def analyze_window(samples, hours, now):
    """找出過去 hours 小時 sbi 都 = TARGET_SBI 且 mday 都相同的站"""
    cutoff = now - timedelta(hours=hours)
    window_samples = [(t, stations) for (t, stations) in samples if t >= cutoff]

    if len(window_samples) < 2:
        # 樣本太少無法判斷（例如剛開始跑，只有 1 次採樣）
        return {"stations": [], "sample_count": len(window_samples), "note": "樣本數不足，無法判斷"}

    # 只考慮「所有樣本都出現過」的站
    station_ids = set(window_samples[0][1].keys())
    for _, stations in window_samples[1:]:
        station_ids &= set(stations.keys())

    hits = []
    for sno in station_ids:
        series_sbi = [stations[sno]["sbi"] for (_, stations) in window_samples]
        series_mday = [stations[sno]["mday"] for (_, stations) in window_samples]

        # 條件 1：所有樣本 sbi 都等於目標值
        if not all(v == TARGET_SBI for v in series_sbi):
            continue
        # 條件 2：所有樣本 mday 都一樣（沒有交易）
        if len(set(series_mday)) != 1:
            continue
        # 過關
        latest = window_samples[-1][1][sno]
        hits.append({
            "sno": sno,
            "sna": latest["sna"],
            "sarea": latest["sarea"],
            "city": latest["city"],
            "sbi": latest["sbi"],
            "bemp": latest["bemp"],
            "tot": latest["tot"],
            "last_mday": latest["mday"],
            "sample_count": len(window_samples),
        })

    # 依城市 → 行政區 → 站名排序
    hits.sort(key=lambda x: (x["city"], x["sarea"], x["sna"]))

    return {
        "stations": hits,
        "sample_count": len(window_samples),
    }


def format_mday(s):
    """把 20260701093300 或 ISO 字串格式化成 MM/DD HH:MM"""
    if not s:
        return ""
    if len(s) == 14 and s.isdigit():
        return f"{s[4:6]}/{s[6:8]} {s[8:10]}:{s[10:12]}"
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d.strftime("%m/%d %H:%M")
    except Exception:
        return s


def write_json(results, now):
    """寫給 HTML 讀的 JSON"""
    payload = {
        "updated_at": now.isoformat(timespec="seconds"),
        "updated_at_display": now.strftime("%Y-%m-%d %H:%M:%S"),
        "target_sbi": TARGET_SBI,
        "windows": {f"{h}h": results[h] for h in WINDOWS},
    }
    with open("stale.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print("✓ 已寫入 stale.json")


def write_excel(results, now):
    """產出 Excel：每個時間區間一個工作表"""
    wb = Workbook()
    wb.remove(wb.active)   # 移除預設工作表

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2A6F97", end_color="2A6F97", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for h in WINDOWS:
        ws = wb.create_sheet(title=f"過去{h}小時只有{TARGET_SBI}台車")
        hits = results[h]["stations"]

        # 頂端說明
        ws["A1"] = f"分析時間：{now.strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"] = f"條件：過去 {h} 小時內採樣，車輛數持續 = {TARGET_SBI} 且無借還交易"
        ws["A3"] = f"樣本數：{results[h]['sample_count']} 次  |  符合站數：{len(hits)}"

        # 標題列
        headers = ["城市", "行政區", "場站名稱", "編號", "最後借還時間", "在站", "空位", "總車格"]
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # 資料
        for row_idx, s in enumerate(hits, start=6):
            ws.cell(row=row_idx, column=1, value=s["city"]).alignment = center
            ws.cell(row=row_idx, column=2, value=s["sarea"]).alignment = center
            ws.cell(row=row_idx, column=3, value=s["sna"]).alignment = left
            ws.cell(row=row_idx, column=4, value=s["sno"]).alignment = left
            ws.cell(row=row_idx, column=5, value=format_mday(s["last_mday"])).alignment = center
            ws.cell(row=row_idx, column=6, value=s["sbi"]).alignment = center
            ws.cell(row=row_idx, column=7, value=s["bemp"]).alignment = center
            ws.cell(row=row_idx, column=8, value=s["tot"]).alignment = center

        # 欄寬
        widths = [8, 12, 32, 20, 16, 8, 8, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        # 凍結標題列
        ws.freeze_panes = "A6"

        # 沒資料時
        if not hits:
            ws.cell(row=6, column=1, value="(目前沒有符合的站點)").alignment = center

    wb.save("stale.xlsx")
    print("✓ 已寫入 stale.xlsx")


def main():
    now = datetime.now(TPE)
    print(f"=== 分析時間：{now.strftime('%Y-%m-%d %H:%M:%S')} ===")
    samples = load_history()
    print(f"讀到 {len(samples)} 個歷史採樣檔")

    results = {}
    for h in WINDOWS:
        result = analyze_window(samples, h, now)
        results[h] = result
        print(f"  過去 {h}h：樣本 {result['sample_count']} 次，符合 {len(result['stations'])} 站")

    write_json(results, now)
    write_excel(results, now)


if __name__ == "__main__":
    main()
